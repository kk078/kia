#!/usr/bin/env python3
"""Ingest structured data (CSV / Excel) into KIA, one clean record per row.

Each row becomes a readable record like:
  TPA Name: Acme Health | Address: 1 Main St | Specialties: Dental, Vision | Status: Active
which preserves the columns so KIA can answer questions about specific fields.

Server-side duplicate detection (content hash) means re-running skips rows already
learned. Use --group N to pack N rows per chunk (fewer, denser records).

Setup:  pip install openpyxl   (only needed for .xlsx)
Usage (PowerShell):
  python scripts/ingest_structured.py C:\\data\\tpa_database.csv
  python scripts/ingest_structured.py C:\\data\\tpa.xlsx --sheet Sheet1 --group 5
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import urllib.request

API = "http://localhost:8000"


def rows_from_csv(path: str) -> tuple[list[str], list[list[str]]]:
    with open(path, newline="", encoding="utf-8-sig", errors="ignore") as f:
        reader = list(csv.reader(f))
    if not reader:
        return [], []
    return reader[0], reader[1:]


def rows_from_xlsx(path: str, sheet: str | None) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("Install the Excel reader first:  pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    data = [[("" if c is None else str(c)) for c in row]
            for row in ws.iter_rows(values_only=True)]
    if not data:
        return [], []
    return data[0], data[1:]


def is_blank(row: list[str]) -> bool:
    return all((c or "").strip() == "" for c in row)


def row_to_record(header: list[str], row: list[str]) -> str:
    parts = []
    for i, cell in enumerate(row):
        col = header[i] if i < len(header) else f"col{i}"
        val = (cell or "").strip()
        if val:
            parts.append(f"{col}: {val}")
    return " | ".join(parts)


def post(api: str, content: str, source: str) -> int:
    body = json.dumps({"content": content, "source": source}).encode()
    req = urllib.request.Request(api.rstrip("/") + "/api/v1/knowledge/ingest",
                                 data=body, headers={"Content-Type": "application/json"},
                                 method="POST")
    with urllib.request.urlopen(req, timeout=600) as r:
        return len(json.loads(r.read().decode()).get("chunk_ids", []))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--api", default=API)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--group", type=int, default=1, help="rows per record (default 1)")
    args = ap.parse_args()

    ext = os.path.splitext(args.file)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        header, rows = rows_from_xlsx(args.file, args.sheet)
    else:
        header, rows = rows_from_csv(args.file)

    label = os.path.basename(args.file)
    print(f"{label}: {len(header)} columns, {len(rows)} data rows")
    if not header:
        raise SystemExit("No header row found.")

    records = [row_to_record(header, r) for r in rows if not is_blank(r)]
    records = [r for r in records if r.strip()]
    print(f"{len(records)} non-empty records")

    indexed = skipped = 0
    buf: list[str] = []
    sent = 0
    for rec in records:
        buf.append(rec)
        if len(buf) >= args.group:
            sent += 1
            content = f"Records from {label}:\n" + "\n".join(buf)
            n = post(args.api, content, f"data:{label}#{sent}")
            if n:
                indexed += 1
            else:
                skipped += 1  # server deduped (already learned)
            print(f"  record {sent}: {'indexed' if n else 'duplicate-skipped'}")
            buf = []
    if buf:
        sent += 1
        content = f"Records from {label}:\n" + "\n".join(buf)
        n = post(args.api, content, f"data:{label}#{sent}")
        indexed += 1 if n else 0
        skipped += 0 if n else 1
        print(f"  record {sent}: {'indexed' if n else 'duplicate-skipped'}")

    print(f"\nDone. records_sent={sent} indexed={indexed} duplicate_skipped={skipped}")
    print("Ask KIA about it in chat with:  /brain <your question>")


if __name__ == "__main__":
    main()
